import os
import sys
import io
import tempfile
import types
import unittest
from unittest import mock
from datetime import datetime, timedelta, timezone

from flask import Flask
from openpyxl import load_workbook


BACKEND_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BACKEND_DIR not in sys.path:
    sys.path.insert(0, BACKEND_DIR)

if "flask_migrate" not in sys.modules:
    flask_migrate = types.ModuleType("flask_migrate")

    class _Migrate:
        def init_app(self, *args, **kwargs):
            return None

    flask_migrate.Migrate = _Migrate
    sys.modules["flask_migrate"] = flask_migrate

if "pythonjsonlogger" not in sys.modules:
    pythonjsonlogger = types.ModuleType("pythonjsonlogger")
    jsonlogger = types.ModuleType("jsonlogger")

    class _JsonFormatter:
        def __init__(self, *args, **kwargs):
            pass

    jsonlogger.JsonFormatter = _JsonFormatter
    pythonjsonlogger.jsonlogger = jsonlogger
    sys.modules["pythonjsonlogger"] = pythonjsonlogger

if "redis" not in sys.modules:
    redis = types.ModuleType("redis")
    redis.from_url = lambda *args, **kwargs: object()
    sys.modules["redis"] = redis

try:
    import chardet  # noqa: F401
except ImportError:
    chardet = types.ModuleType("chardet")
    chardet.detect = lambda content: {"encoding": "utf-8"}
    sys.modules["chardet"] = chardet

from app import db  # noqa: E402
import app.models  # noqa: F401,E402
from app.api.consumption import bp as consumption_bp  # noqa: E402
from app.models import (  # noqa: E402
    ConsumptionRecord,
    ConsumptionSyncState,
    TimeCalibrationSample,
    MatchResult,
    MatchStatusEnum,
    CapturedImage,
    Dish,
    DishRecognition,
    Student,
    CategoryEnum,
    ImageStatusEnum,
    RoleEnum,
    User,
)
from app.services.consumption_location_filter import ENABLED_TRANSACTION_LOCATION_IDS_KEY  # noqa: E402
from app.utils.jwt_utils import generate_token  # noqa: E402


class ConsumptionApiTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        runtime_config = tempfile.NamedTemporaryFile(delete=False)
        runtime_config.close()
        cls.runtime_config_path = runtime_config.name
        cls.app = Flask(__name__)
        cls.app.config.update(
            SQLALCHEMY_DATABASE_URI="sqlite:///:memory:",
            SQLALCHEMY_TRACK_MODIFICATIONS=False,
            SECRET_KEY="test-secret",
            JWT_ALGORITHM="HS256",
            JWT_ACCESS_TOKEN_EXPIRES=timedelta(hours=1),
            LOCAL_RUNTIME_CONFIG_PATH=cls.runtime_config_path,
        )
        db.init_app(cls.app)
        cls.app.register_blueprint(consumption_bp, url_prefix="/api/v1/consumption")
        cls.app_context = cls.app.app_context()
        cls.app_context.push()
        db.create_all()
        cls.client = cls.app.test_client()

    @classmethod
    def tearDownClass(cls):
        db.session.remove()
        db.drop_all()
        cls.app_context.pop()
        try:
            os.unlink(cls.runtime_config_path)
        except OSError:
            pass

    def setUp(self):
        with open(self.runtime_config_path, "w", encoding="utf-8") as f:
            f.write("{}")
        self.app.config.update(
            ZTK_SYNC_ENABLED=False,
            ZTK_DB_HOST="",
            ZTK_DB_NAME="ZYTK40_PLUS",
            ZTK_DB_USER="",
            ZTK_DB_PASSWORD="",
            ZTK_SYNC_INTERVAL_MINUTES=5,
            TIME_OFFSET_CALIBRATION=0.0,
            **{ENABLED_TRANSACTION_LOCATION_IDS_KEY: []},
        )
        db.session.query(MatchResult).delete()
        db.session.query(DishRecognition).delete()
        db.session.query(CapturedImage).delete()
        db.session.query(Dish).delete()
        db.session.query(ConsumptionRecord).delete()
        db.session.query(ConsumptionSyncState).delete()
        db.session.query(TimeCalibrationSample).delete()
        db.session.query(Student).delete()
        db.session.query(User).delete()
        db.session.commit()

        admin = User(
            username="admin",
            name="管理员",
            role=RoleEnum.admin,
            is_active=True,
        )
        db.session.add(admin)
        db.session.commit()
        self.admin_id = admin.id

    def tearDown(self):
        db.session.rollback()

    def _auth_headers(self) -> dict[str, str]:
        token = generate_token(self.admin_id, RoleEnum.admin.value)
        return {"Authorization": f"Bearer {token}"}

    def test_list_matches_is_consumption_record_driven(self):
        record = ConsumptionRecord(
            student_no="230501",
            student_name="张三",
            transaction_time=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc),
            amount=-12.0,
            transaction_id="tx-001",
        )
        db.session.add(record)
        db.session.commit()

        res = self.client.get(
            "/api/v1/consumption/matches?date=2026-03-31",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        self.assertEqual(payload["data"]["total"], 1)

        item = payload["data"]["items"][0]
        self.assertEqual(item["consumption_record_id"], record.id)
        self.assertEqual(item["status"], "unmatched_record")
        self.assertEqual(item["consumption_record"]["transaction_id"], "tx-001")

    def test_list_matches_all_includes_matched_and_unmatched_consumption_records(self):
        matched_record = ConsumptionRecord(
            student_no="230501",
            student_name="张三",
            transaction_time=datetime(2026, 3, 31, 12, 5, tzinfo=timezone.utc),
            amount=-12.0,
            transaction_id="tx-all-matched",
        )
        unmatched_record = ConsumptionRecord(
            student_no="230502",
            student_name="李四",
            transaction_time=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc),
            amount=-8.0,
            transaction_id="tx-all-unmatched",
        )
        db.session.add_all([matched_record, unmatched_record])
        db.session.flush()
        db.session.add(MatchResult(
            consumption_record_id=matched_record.id,
            status=MatchStatusEnum.matched,
            match_date=matched_record.transaction_time.date(),
        ))
        db.session.commit()

        res = self.client.get(
            "/api/v1/consumption/matches?date=2026-03-31",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        self.assertEqual(payload["data"]["total"], 2)
        statuses_by_transaction = {
            item["consumption_record"]["transaction_id"]: item["status"]
            for item in payload["data"]["items"]
        }
        self.assertEqual(statuses_by_transaction["tx-all-matched"], "matched")
        self.assertEqual(statuses_by_transaction["tx-all-unmatched"], "unmatched_record")

    def test_list_matches_returns_status_counts_ignoring_status_filter(self):
        def make_record(tx_id, hour):
            record = ConsumptionRecord(
                student_no="230501",
                student_name="张三",
                transaction_time=datetime(2026, 3, 31, hour, 0, tzinfo=timezone.utc),
                amount=-10.0,
                transaction_id=tx_id,
            )
            db.session.add(record)
            db.session.flush()
            return record

        matched_record = make_record("tx-sc-matched", 11)
        pending_record = make_record("tx-sc-pending", 12)
        confirmed_record = make_record("tx-sc-confirmed", 13)
        unmatched_match_record = make_record("tx-sc-unmatched-match", 14)
        make_record("tx-sc-no-match", 15)
        db.session.add_all([
            MatchResult(
                consumption_record_id=matched_record.id,
                status=MatchStatusEnum.matched,
                match_date=matched_record.transaction_time.date(),
            ),
            MatchResult(
                consumption_record_id=pending_record.id,
                status=MatchStatusEnum.time_matched_only,
                match_date=pending_record.transaction_time.date(),
            ),
            MatchResult(
                consumption_record_id=confirmed_record.id,
                status=MatchStatusEnum.confirmed,
                match_date=confirmed_record.transaction_time.date(),
            ),
            MatchResult(
                consumption_record_id=unmatched_match_record.id,
                status=MatchStatusEnum.unmatched_record,
                match_date=unmatched_match_record.transaction_time.date(),
            ),
        ])
        db.session.commit()

        for query in ("date=2026-03-31", "date=2026-03-31&status=matched"):
            res = self.client.get(
                f"/api/v1/consumption/matches?{query}",
                headers=self._auth_headers(),
            )
            self.assertEqual(res.status_code, 200)
            status_counts = res.get_json()["data"]["status_counts"]
            self.assertEqual(status_counts["total"], 5)
            self.assertEqual(status_counts["matched"], 1)
            self.assertEqual(status_counts["time_matched_only"], 1)
            self.assertEqual(status_counts["confirmed"], 1)
            # 无 MatchResult 的记录 + MatchResult 状态为 unmatched_record 的记录
            self.assertEqual(status_counts["unmatched_record"], 2)

    def test_list_matches_excludes_positive_recharge_records(self):
        consumption_record = ConsumptionRecord(
            student_no="230501",
            student_name="张三",
            transaction_time=datetime(2026, 3, 31, 12, 5, tzinfo=timezone.utc),
            amount=-12.0,
            transaction_id="tx-negative-consumption",
        )
        recharge_record = ConsumptionRecord(
            student_no="230502",
            student_name="李四",
            transaction_time=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc),
            amount=20.0,
            transaction_id="tx-positive-recharge",
        )
        db.session.add_all([consumption_record, recharge_record])
        db.session.commit()

        res = self.client.get(
            "/api/v1/consumption/matches?date=2026-03-31",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        self.assertEqual(payload["data"]["total"], 1)
        self.assertEqual(
            payload["data"]["items"][0]["consumption_record"]["transaction_id"],
            "tx-negative-consumption",
        )

    def test_list_matches_filters_by_enabled_transaction_location_ids(self):
        self.app.config[ENABLED_TRANSACTION_LOCATION_IDS_KEY] = ["1-15"]
        db.session.add_all([
            ConsumptionRecord(
                student_no="230501",
                student_name="张三",
                transaction_time=datetime(2026, 3, 31, 12, 5, tzinfo=timezone.utc),
                amount=-12.0,
                transaction_id="tx-match-enabled-location",
                channel_id="1-15",
            ),
            ConsumptionRecord(
                student_no="230502",
                student_name="李四",
                transaction_time=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc),
                amount=-8.0,
                transaction_id="tx-match-disabled-location",
                channel_id="1-16",
            ),
        ])
        db.session.commit()

        res = self.client.get(
            "/api/v1/consumption/matches?date=2026-03-31",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        self.assertEqual(payload["data"]["total"], 1)
        self.assertEqual(
            payload["data"]["items"][0]["consumption_record"]["transaction_id"],
            "tx-match-enabled-location",
        )

    def test_list_matches_all_paginates_consumption_records_not_match_rows(self):
        matched_record = ConsumptionRecord(
            student_no="230501",
            student_name="张三",
            transaction_time=datetime(2026, 3, 31, 12, 5, tzinfo=timezone.utc),
            amount=-12.0,
            transaction_id="tx-duplicate-match-record",
        )
        unmatched_record = ConsumptionRecord(
            student_no="230502",
            student_name="李四",
            transaction_time=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc),
            amount=-8.0,
            transaction_id="tx-page-unmatched-record",
        )
        db.session.add_all([matched_record, unmatched_record])
        db.session.flush()
        db.session.add_all([
            MatchResult(
                consumption_record_id=matched_record.id,
                status=MatchStatusEnum.matched,
                match_date=matched_record.transaction_time.date(),
            ),
            MatchResult(
                consumption_record_id=matched_record.id,
                status=MatchStatusEnum.time_matched_only,
                match_date=matched_record.transaction_time.date(),
            ),
        ])
        db.session.commit()

        res = self.client.get(
            "/api/v1/consumption/matches?date=2026-03-31&page_size=2",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        self.assertEqual(payload["data"]["total"], 2)
        self.assertCountEqual(
            [item["consumption_record"]["transaction_id"] for item in payload["data"]["items"]],
            ["tx-duplicate-match-record", "tx-page-unmatched-record"],
        )

    def test_import_uses_time_based_batch_id(self):
        content = (
            "学号,学生姓名,消费时间,消费金额,流水号,交易地点\n"
            "230501,张三,2026-03-31 12:05:30,12.50,TX202603310001,1\n"
        ).encode("gbk")

        fake_matching = types.ModuleType("app.tasks.matching")
        delay_mock = mock.Mock()
        fake_matching.run_matching_for_batch = types.SimpleNamespace(delay=delay_mock)

        with mock.patch.dict(sys.modules, {"app.tasks.matching": fake_matching}):
            res = self.client.post(
                "/api/v1/consumption/import",
                data={"file": (io.BytesIO(content), "records.csv")},
                content_type="multipart/form-data",
                headers=self._auth_headers(),
            )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        batch_id = payload["data"]["batch_id"]
        self.assertRegex(batch_id, r"^\d{17}$")
        self.assertEqual(ConsumptionRecord.query.one().import_batch, batch_id)
        self.assertEqual(ConsumptionRecord.query.one().channel_id, "1")
        delay_mock.assert_called_once_with(batch_id)

    def test_db_sync_status_returns_state(self):
        self.app.config.update(
            ZTK_SYNC_ENABLED=True,
            ZTK_DB_HOST="sqlserver.example.local",
            ZTK_DB_NAME="ZYTK40_PLUS",
            ZTK_DB_USER="test-user",
            ZTK_DB_PASSWORD="test-password",
        )
        db.session.add(ConsumptionSyncState(
            source_system="ztk_plus",
            cursor_source_record_id="1001",
            last_batch_id="ztk-test-batch",
            last_success_count=2,
        ))
        db.session.commit()

        res = self.client.get(
            "/api/v1/consumption/db-sync/status",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        data = payload["data"]
        self.assertTrue(data["enabled"])
        self.assertEqual(data["sync_interval_minutes"], 5)
        self.assertTrue(data["configured"])
        self.assertEqual(data["state"]["cursor_source_record_id"], "1001")
        self.assertEqual(data["state"]["last_batch_id"], "ztk-test-batch")

    def test_db_sync_config_can_update_enable_and_interval(self):
        res = self.client.put(
            "/api/v1/consumption/db-sync/config",
            json={
                "host": "sqlserver.example.local",
                "port": "1433",
                "database": "ZYTK40_PLUS",
                "user": "test-user",
                "password": "test-password",
                "payment_books_table": "dbo.view_ac_paymentbooks",
                "sync_enabled": True,
                "sync_interval_minutes": "12",
                "enabled_transaction_location_ids": "1-15\n1-16, 1-15",
            },
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        self.assertTrue(payload["data"]["sync_enabled"])
        self.assertEqual(payload["data"]["sync_interval_minutes"], 12)
        self.assertTrue(payload["data"]["has_password"])
        self.assertEqual(payload["data"]["enabled_transaction_location_ids"], ["1-15", "1-16"])

        res = self.client.get(
            "/api/v1/consumption/db-sync/config",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        self.assertTrue(payload["data"]["sync_enabled"])
        self.assertEqual(payload["data"]["sync_interval_minutes"], 12)
        self.assertEqual(payload["data"]["enabled_transaction_location_ids"], ["1-15", "1-16"])

    def test_db_sync_trigger_submits_forced_task(self):
        fake_task_module = types.ModuleType("app.tasks.ztk_consumption")
        delay_mock = mock.Mock()
        fake_task_module.sync_ztk_consumption = types.SimpleNamespace(delay=delay_mock)

        with mock.patch.dict(sys.modules, {"app.tasks.ztk_consumption": fake_task_module}):
            res = self.client.post(
                "/api/v1/consumption/db-sync/trigger",
                headers=self._auth_headers(),
            )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        self.assertIn("已提交", payload["data"]["message"])
        delay_mock.assert_called_once_with(True)

    def test_list_records_filters_by_import_batch(self):
        db.session.add_all([
            ConsumptionRecord(
                student_no="230501",
                student_name="张三",
                transaction_time=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc),
                amount=12.0,
                transaction_id="tx-batch-001",
                import_batch="20260331120000001",
            ),
            ConsumptionRecord(
                student_no="230502",
                student_name="李四",
                transaction_time=datetime(2026, 3, 31, 12, 5, tzinfo=timezone.utc),
                amount=8.0,
                transaction_id="tx-batch-002",
                import_batch="20260331120500001",
            ),
        ])
        db.session.commit()

        res = self.client.get(
            "/api/v1/consumption/records?batch=20260331120000001",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        self.assertEqual(payload["data"]["total"], 1)
        self.assertEqual(payload["data"]["items"][0]["transaction_id"], "tx-batch-001")

    def test_list_records_filters_by_date_range_and_text_fields(self):
        db.session.add_all([
            ConsumptionRecord(
                student_no="230501",
                student_name="张三",
                transaction_time=datetime(2026, 3, 30, 12, 0, tzinfo=timezone.utc),
                amount=10.0,
                transaction_id="tx-filter-out-date",
                channel_id="1",
                import_batch="20260330120000001",
            ),
            ConsumptionRecord(
                student_no="230502",
                student_name="张三丰",
                transaction_time=datetime(2026, 3, 31, 12, 5, tzinfo=timezone.utc),
                amount=12.0,
                transaction_id="tx-filter-hit-001",
                channel_id="一食堂一楼",
                import_batch="20260331120500001",
            ),
            ConsumptionRecord(
                student_no="230503",
                student_name="李四",
                transaction_time=datetime(2026, 3, 31, 12, 10, tzinfo=timezone.utc),
                amount=8.0,
                transaction_id="tx-filter-out-student",
                channel_id="一食堂一楼",
                import_batch="20260331121000001",
            ),
        ])
        db.session.commit()

        res = self.client.get(
            "/api/v1/consumption/records?date_from=2026-03-31&date_to=2026-03-31&student=张三&channel_id=一食堂&transaction_id=hit",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        self.assertEqual(payload["data"]["total"], 1)
        self.assertEqual(payload["data"]["items"][0]["transaction_id"], "tx-filter-hit-001")

    def test_student_number_filter_does_not_include_card_number_matches(self):
        student = Student(
            student_no="20260001",
            name="张三",
            card_no="C1001",
            is_active=True,
        )
        db.session.add(student)
        db.session.flush()
        db.session.add_all([
            ConsumptionRecord(
                student_no="20269999",
                card_code="C1001",
                student_name="原始卡号记录",
                transaction_time=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc),
                amount=-8.0,
                transaction_id="tx-card-number",
            ),
            ConsumptionRecord(
                student_id=student.id,
                student_no="20260001",
                student_name="张三",
                transaction_time=datetime(2026, 3, 31, 12, 5, tzinfo=timezone.utc),
                amount=-12.0,
                transaction_id="tx-card-in-source-payload",
                source_payload={"CardCode": "C1001"},
            ),
            ConsumptionRecord(
                student_id=student.id,
                student_no="20260001",
                student_name="张三",
                transaction_time=datetime(2026, 3, 31, 12, 10, tzinfo=timezone.utc),
                amount=-10.0,
                transaction_id="tx-internal-id-only",
            ),
            ConsumptionRecord(
                student_no="C9999",
                student_name="张三",
                transaction_time=datetime(2026, 3, 31, 12, 15, tzinfo=timezone.utc),
                amount=-6.0,
                transaction_id="tx-same-name-only",
            ),
        ])
        db.session.commit()

        records_res = self.client.get(
            "/api/v1/consumption/records?student_no=20260001",
            headers=self._auth_headers(),
        )
        matches_res = self.client.get(
            "/api/v1/consumption/matches?student_no=20260001",
            headers=self._auth_headers(),
        )

        self.assertEqual(records_res.status_code, 200)
        self.assertEqual(matches_res.status_code, 200)
        self.assertCountEqual(
            [item["transaction_id"] for item in records_res.get_json()["data"]["items"]],
            ["tx-card-in-source-payload", "tx-internal-id-only"],
        )
        self.assertCountEqual(
            [
                item["consumption_record"]["transaction_id"]
                for item in matches_res.get_json()["data"]["items"]
            ],
            ["tx-card-in-source-payload", "tx-internal-id-only"],
        )

    def test_student_number_filter_resolves_ztk_acc_num_without_card_number(self):
        db.session.add(Student(
            student_no="20260002",
            name="李四",
            is_active=True,
        ))
        db.session.add(ConsumptionRecord(
            student_no="20260002",
            card_code="C1002",
            student_name="李四",
            transaction_time=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc),
            amount=-8.0,
            transaction_id="tx-acc-num",
            source_system="ztk_plus",
            source_payload={"AccNum": 20260002, "CardCode": "C1002"},
        ))
        db.session.commit()

        res = self.client.get(
            "/api/v1/consumption/records?student_no=20260002",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        self.assertEqual(
            [item["transaction_id"] for item in res.get_json()["data"]["items"]],
            ["tx-acc-num"],
        )
        self.assertEqual(res.get_json()["data"]["items"][0]["card_code"], "C1002")

    def test_list_records_filters_by_enabled_transaction_location_ids(self):
        self.app.config[ENABLED_TRANSACTION_LOCATION_IDS_KEY] = ["1-15"]
        db.session.add_all([
            ConsumptionRecord(
                student_no="230501",
                student_name="张三",
                transaction_time=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc),
                amount=-12.0,
                transaction_id="tx-enabled-location",
                channel_id="1-15",
            ),
            ConsumptionRecord(
                student_no="230502",
                student_name="李四",
                transaction_time=datetime(2026, 3, 31, 12, 5, tzinfo=timezone.utc),
                amount=-8.0,
                transaction_id="tx-disabled-location",
                channel_id="1-16",
            ),
        ])
        db.session.commit()

        res = self.client.get(
            "/api/v1/consumption/records?date=2026-03-31",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        self.assertEqual(payload["data"]["total"], 1)
        self.assertEqual(payload["data"]["items"][0]["transaction_id"], "tx-enabled-location")

    def test_list_records_rejects_invalid_date_range(self):
        res = self.client.get(
            "/api/v1/consumption/records?date_from=2026-04-01&date_to=2026-03-31",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 400)
        payload = res.get_json()
        self.assertEqual(payload["code"], 400)
        self.assertEqual(payload["message"], "开始日期不能晚于结束日期")

    def test_list_records_does_not_expose_raw_source_payload(self):
        db.session.add(ConsumptionRecord(
            student_no="230501",
            student_name="张三",
            transaction_time=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc),
            amount=12.0,
            transaction_id="ztk:PaymentBooks:1001",
            source_system="ztk_plus",
            source_record_id="1001",
            source_payload={
                "AccNum": 80000001,
                "CardCode": "000123",
                "MonDBCurr": 100.50,
                "RecTime": "2026-03-31T12:00:03",
            },
        ))
        db.session.commit()

        res = self.client.get(
            "/api/v1/consumption/records",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        item = res.get_json()["data"]["items"][0]
        self.assertEqual(item["source_system"], "ztk_plus")
        self.assertEqual(item["source_record_id"], "1001")
        self.assertEqual(item["rec_time"], "2026-03-31T12:00:03")
        self.assertNotIn("source_payload", item)

    def test_get_record_returns_same_minute_time_calibration(self):
        record = ConsumptionRecord(
            student_no="230501",
            student_name="张三",
            transaction_time=datetime(2026, 3, 31, 12, 0, 40),
            amount=12.0,
            transaction_id="tx-calibration-detail",
            source_system="ztk_plus",
        )
        sample = TimeCalibrationSample(
            source_system="ztk_plus",
            source_time=datetime(2026, 3, 31, 12, 0, 5),
            local_time=datetime(2026, 3, 31, 12, 0, 2, 750000),
            offset_seconds=2.25,
            rtt_ms=4.5,
        )
        db.session.add_all([record, sample])
        db.session.commit()

        res = self.client.get(
            f"/api/v1/consumption/records/{record.id}",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        data = res.get_json()["data"]
        self.assertEqual(data["record"]["transaction_id"], "tx-calibration-detail")
        calibration = data["time_calibration"]
        self.assertEqual(calibration["resolution_method"], "same_minute")
        self.assertEqual(calibration["offset_seconds"], 2.25)
        self.assertEqual(calibration["adjustment_seconds"], -2.25)
        self.assertEqual(calibration["sample_distance_seconds"], 35.0)
        self.assertEqual(calibration["rtt_ms"], 4.5)
        self.assertEqual(calibration["aligned_transaction_time"], "2026-03-31T12:00:37.750000")

    def test_get_record_returns_manual_fallback_without_samples(self):
        self.app.config["TIME_OFFSET_CALIBRATION"] = -1.5
        record = ConsumptionRecord(
            transaction_time=datetime(2026, 3, 31, 12, 0),
            amount=8.0,
            transaction_id="tx-calibration-fallback",
        )
        db.session.add(record)
        db.session.commit()

        res = self.client.get(
            f"/api/v1/consumption/records/{record.id}",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        calibration = res.get_json()["data"]["time_calibration"]
        self.assertEqual(calibration["resolution_method"], "manual_fallback")
        self.assertEqual(calibration["offset_seconds"], -1.5)
        self.assertEqual(calibration["adjustment_seconds"], -1.5)
        self.assertIsNone(calibration["source_time"])
        self.assertIsNone(calibration["sample_distance_seconds"])

    def test_list_record_batches_groups_imports(self):
        db.session.add_all([
            ConsumptionRecord(
                student_no="230501",
                student_name="张三",
                transaction_time=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc),
                amount=12.0,
                transaction_id="tx-group-001",
                import_batch="20260331120000001",
            ),
            ConsumptionRecord(
                student_no="230502",
                student_name="李四",
                transaction_time=datetime(2026, 3, 31, 12, 5, tzinfo=timezone.utc),
                amount=8.0,
                transaction_id="tx-group-002",
                import_batch="20260331120000001",
            ),
            ConsumptionRecord(
                student_no="230503",
                student_name="王五",
                transaction_time=datetime(2026, 3, 31, 12, 10, tzinfo=timezone.utc),
                amount=10.0,
                transaction_id="tx-group-003",
                import_batch="20260331121000001",
            ),
        ])
        db.session.commit()

        res = self.client.get(
            "/api/v1/consumption/records/batches",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        batches = {item["batch_id"]: item for item in payload["data"]["items"]}
        self.assertEqual(batches["20260331120000001"]["record_count"], 2)
        self.assertEqual(batches["20260331120000001"]["total_amount"], 20.0)
        self.assertEqual(batches["20260331121000001"]["record_count"], 1)

    def test_delete_record_removes_related_match(self):
        record = ConsumptionRecord(
            student_no="230501",
            student_name="张三",
            transaction_time=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc),
            amount=12.0,
            transaction_id="tx-delete-001",
            import_batch="20260331120000001",
        )
        db.session.add(record)
        db.session.commit()
        db.session.add(MatchResult(
            consumption_record_id=record.id,
            status=MatchStatusEnum.unmatched_record,
            match_date=record.transaction_time.date(),
        ))
        db.session.commit()

        res = self.client.delete(
            f"/api/v1/consumption/records/{record.id}",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        self.assertEqual(payload["data"]["deleted"], 1)
        self.assertEqual(ConsumptionRecord.query.count(), 0)
        self.assertEqual(MatchResult.query.count(), 0)

    def test_delete_batch_removes_all_batch_records_and_matches(self):
        keep = ConsumptionRecord(
            student_no="230501",
            student_name="张三",
            transaction_time=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc),
            amount=12.0,
            transaction_id="tx-keep-001",
            import_batch="20260331120000001",
        )
        delete_one = ConsumptionRecord(
            student_no="230502",
            student_name="李四",
            transaction_time=datetime(2026, 3, 31, 12, 5, tzinfo=timezone.utc),
            amount=8.0,
            transaction_id="tx-delete-batch-001",
            import_batch="20260331120500001",
        )
        delete_two = ConsumptionRecord(
            student_no="230503",
            student_name="王五",
            transaction_time=datetime(2026, 3, 31, 12, 10, tzinfo=timezone.utc),
            amount=10.0,
            transaction_id="tx-delete-batch-002",
            import_batch="20260331120500001",
        )
        db.session.add_all([keep, delete_one, delete_two])
        db.session.commit()
        db.session.add_all([
            MatchResult(
                consumption_record_id=delete_one.id,
                status=MatchStatusEnum.unmatched_record,
                match_date=delete_one.transaction_time.date(),
            ),
            MatchResult(
                consumption_record_id=delete_two.id,
                status=MatchStatusEnum.unmatched_record,
                match_date=delete_two.transaction_time.date(),
            ),
        ])
        db.session.commit()

        res = self.client.delete(
            "/api/v1/consumption/records/batches/20260331120500001",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        self.assertEqual(payload["data"]["deleted"], 2)
        self.assertEqual(ConsumptionRecord.query.count(), 1)
        self.assertEqual(ConsumptionRecord.query.one().transaction_id, "tx-keep-001")
        self.assertEqual(MatchResult.query.count(), 0)

    def test_import_settings_can_be_updated(self):
        res = self.client.put(
            "/api/v1/consumption/import-settings",
            json={"allowed_locations": ["1", " 2 "]},
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        self.assertEqual(payload["data"]["allowed_locations"], ["1", "2"])

        res = self.client.get(
            "/api/v1/consumption/import-settings",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        self.assertEqual(payload["data"]["allowed_locations"], ["1", "2"])

    def test_download_import_template_returns_excel_file(self):
        res = self.client.get(
            "/api/v1/consumption/import-template",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        self.assertEqual(
            res.mimetype,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment", res.headers.get("Content-Disposition", ""))

        workbook = load_workbook(io.BytesIO(res.data))
        sheet = workbook.active
        self.assertEqual(sheet.title, "消费记录导入模板")
        self.assertEqual(
            [sheet.cell(row=1, column=col).value for col in range(1, 7)],
            ["学号/消费卡号 *", "学生姓名", "消费时间 *", "消费金额 *", "流水号 *", "通道 *"],
        )
        self.assertEqual(sheet.cell(row=2, column=1).value, "230501")
        self.assertEqual(sheet.cell(row=4, column=1).value[:3], "说明：")

    def test_list_unmatched_images_returns_image_payload(self):
        image = CapturedImage(
            capture_date=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc).date(),
            channel_id="1",
            captured_at=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc),
            image_path="/tmp/unmatched.jpg",
            status=ImageStatusEnum.identified,
            source_video="nvr_001.mp4",
            is_candidate=False,
        )
        dish = Dish(
            name="青椒土豆丝",
            price=6.0,
            category=CategoryEnum.vegetable,
            is_active=True,
        )
        db.session.add(dish)
        db.session.add(image)
        db.session.commit()

        recognition = DishRecognition(
            image_id=image.id,
            dish_id=dish.id,
            dish_name_raw=dish.name,
            confidence=0.91,
            is_low_confidence=False,
            is_manual=False,
            model_version="test",
        )

        match = MatchResult(
            image_id=image.id,
            status=MatchStatusEnum.unmatched_image,
            match_date=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc).date(),
        )
        db.session.add(recognition)
        db.session.add(match)
        db.session.commit()

        res = self.client.get(
            "/api/v1/consumption/matches/unmatched-images?date=2026-03-31",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        self.assertEqual(payload["data"]["total"], 1)
        item = payload["data"]["items"][0]
        self.assertEqual(item["status"], "unmatched_image")
        self.assertEqual(item["captured_at"], image.captured_at.isoformat())
        self.assertEqual(item["image"]["id"], image.id)
        self.assertEqual(item["image"]["source_video"], "nvr_001.mp4")
        self.assertEqual(item["image"]["recognitions"][0]["dish_price"], 6.0)
        self.assertEqual(
            item["image"]["recognitions"][0]["captured_at"],
            image.captured_at.isoformat(),
        )

    def test_list_matches_returns_linked_image_payload(self):
        record = ConsumptionRecord(
            student_no="230501",
            student_name="张三",
            transaction_time=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc),
            amount=-12.0,
            transaction_id="tx-002",
        )
        db.session.add(record)
        db.session.commit()

        image = CapturedImage(
            capture_date=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc).date(),
            channel_id="2",
            captured_at=datetime(2026, 3, 31, 12, 0, 1, tzinfo=timezone.utc),
            image_path="/tmp/matched.jpg",
            status=ImageStatusEnum.identified,
            source_video="nvr_002.mp4",
            is_candidate=False,
        )
        db.session.add(image)
        db.session.commit()

        match = MatchResult(
            consumption_record_id=record.id,
            image_id=image.id,
            status=MatchStatusEnum.time_matched_only,
            match_date=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc).date(),
        )
        dish = Dish(
            name="红烧肉",
            price=12.0,
            category=CategoryEnum.meat,
            is_active=True,
        )
        db.session.add(dish)
        db.session.commit()

        recognition = DishRecognition(
            image_id=image.id,
            dish_id=dish.id,
            dish_name_raw=dish.name,
            confidence=0.95,
            is_low_confidence=False,
            is_manual=False,
            model_version="test",
        )
        db.session.add(match)
        db.session.add(recognition)
        db.session.commit()

        res = self.client.get(
            "/api/v1/consumption/matches?date=2026-03-31&status=time_matched_only",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        self.assertEqual(payload["code"], 0)
        self.assertEqual(payload["data"]["total"], 1)
        item = payload["data"]["items"][0]
        self.assertEqual(item["status"], "time_matched_only")
        self.assertEqual(item["image"]["id"], image.id)
        self.assertEqual(item["image"]["channel_id"], "2")
        self.assertEqual(item["image_price_total"], 12.0)
        self.assertEqual(item["image"]["recognitions"][0]["dish_price"], 12.0)

        detail_res = self.client.get(
            f"/api/v1/consumption/matches/{match.id}",
            headers=self._auth_headers(),
        )
        self.assertEqual(detail_res.status_code, 200)
        detail = detail_res.get_json()["data"]
        self.assertEqual(detail["id"], match.id)
        self.assertEqual(detail["consumption_record"]["id"], record.id)
        self.assertEqual(detail["image"]["id"], image.id)
        self.assertEqual(detail["image"]["recognitions"][0]["dish_price"], 12.0)

    def test_confirm_match_copies_capture_time_when_reassigning_image(self):
        first_time = datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc)
        second_time = first_time + timedelta(seconds=17)
        first_image = CapturedImage(
            capture_date=first_time.date(),
            channel_id="1",
            captured_at=first_time,
            image_path="/tmp/confirm-first.jpg",
            status=ImageStatusEnum.identified,
            is_candidate=False,
        )
        second_image = CapturedImage(
            capture_date=second_time.date(),
            channel_id="1",
            captured_at=second_time,
            image_path="/tmp/confirm-second.jpg",
            status=ImageStatusEnum.identified,
            is_candidate=False,
        )
        db.session.add_all([first_image, second_image])
        db.session.flush()
        match = MatchResult(
            image_id=first_image.id,
            captured_at=first_image.captured_at,
            status=MatchStatusEnum.unmatched_image,
            match_date=first_time.date(),
        )
        db.session.add(match)
        db.session.commit()

        res = self.client.put(
            f"/api/v1/consumption/matches/{match.id}/confirm",
            headers=self._auth_headers(),
            json={"image_id": second_image.id},
        )

        self.assertEqual(res.status_code, 200)
        refreshed = db.session.get(MatchResult, match.id)
        self.assertEqual(refreshed.image_id, second_image.id)
        self.assertEqual(refreshed.captured_at, second_image.captured_at)
        self.assertEqual(res.get_json()["data"]["captured_at"], second_image.captured_at.isoformat())

    def test_list_matches_recalculates_price_diff_from_current_recognitions(self):
        record = ConsumptionRecord(
            student_no="230501",
            student_name="张三",
            transaction_time=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc),
            amount=-12.0,
            transaction_id="tx-realtime-price-diff",
        )
        image = CapturedImage(
            capture_date=datetime(2026, 3, 31, 12, 0, tzinfo=timezone.utc).date(),
            channel_id="2",
            captured_at=datetime(2026, 3, 31, 12, 0, 1, tzinfo=timezone.utc),
            image_path="/tmp/realtime-price-diff.jpg",
            status=ImageStatusEnum.identified,
            is_candidate=False,
        )
        db.session.add_all([record, image])
        db.session.flush()
        match = MatchResult(
            consumption_record_id=record.id,
            image_id=image.id,
            status=MatchStatusEnum.time_matched_only,
            match_date=record.transaction_time.date(),
            price_diff=None,
        )
        dish = Dish(
            name="红烧肉",
            price=10.0,
            category=CategoryEnum.meat,
            is_active=True,
        )
        db.session.add_all([match, dish])
        db.session.flush()
        db.session.add(DishRecognition(
            image_id=image.id,
            dish_id=dish.id,
            dish_name_raw=dish.name,
            confidence=0.95,
            is_low_confidence=False,
            is_manual=False,
            model_version="test",
        ))
        db.session.commit()

        res = self.client.get(
            "/api/v1/consumption/matches?date=2026-03-31",
            headers=self._auth_headers(),
        )

        self.assertEqual(res.status_code, 200)
        payload = res.get_json()
        item = payload["data"]["items"][0]
        self.assertEqual(item["image_price_total"], 10.0)
        self.assertEqual(item["price_diff"], 2.0)


if __name__ == "__main__":
    unittest.main()
