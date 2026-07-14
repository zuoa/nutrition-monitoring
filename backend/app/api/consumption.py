import logging
import io
from datetime import date, datetime
from flask import Blueprint, current_app, request, send_file
from sqlalchemy import exists, or_
from sqlalchemy.orm import joinedload, selectinload
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from app import db
from app.models import ConsumptionRecord, MatchResult, MatchStatusEnum, DishRecognition
from app.services.runtime_config import get_effective_config, persist_runtime_overrides
from app.services.ztk_consumption_sync import _normalize_text
from app.utils.jwt_utils import login_required, role_required, api_ok, api_error
from app.utils.pagination import paginate, paginated_response
from app.services.import_service import ConsumptionImportService, normalize_allowed_transaction_locations
from app.services.consumption_location_filter import (
    ENABLED_TRANSACTION_LOCATION_IDS_KEY,
    apply_enabled_transaction_location_filter,
    get_enabled_transaction_location_ids,
)

bp = Blueprint("consumption", __name__)
logger = logging.getLogger(__name__)

CONSUMPTION_ALLOWED_LOCATIONS_KEY = "CONSUMPTION_IMPORT_ALLOWED_LOCATIONS"


def _calc_image_price_total(image_id: int | None) -> float:
    if not image_id:
        return 0.0

    total = 0.0
    recognitions = DishRecognition.query.filter(
        DishRecognition.image_id == image_id,
        DishRecognition.is_low_confidence.is_(False),
    ).all()
    for recognition in recognitions:
        if recognition.dish_id and recognition.dish and recognition.dish.price is not None:
            total += float(recognition.dish.price)
    return total


def _build_match_payload(match: MatchResult, record: ConsumptionRecord | None = None) -> dict:
    payload = match.to_dict()
    image_price_total = _calc_image_price_total(match.image_id)
    payload["image_price_total"] = image_price_total

    matched_record = record or match.consumption_record
    if match.image_id and matched_record and matched_record.amount is not None:
        payload["price_diff"] = abs(abs(float(matched_record.amount)) - image_price_total)

    return payload


def _build_match_detail_payload(match: MatchResult) -> dict:
    record = match.consumption_record
    payload = _build_match_payload(match, record)
    if record:
        payload["consumption_record"] = record.to_dict()
    if match.student:
        payload["student"] = match.student.to_dict()
    if match.image:
        image = match.image.to_dict()
        recognitions = DishRecognition.query.filter_by(image_id=match.image.id).all()
        image["recognitions"] = [recognition.to_dict() for recognition in recognitions]
        payload["image"] = image
    return payload


def _match_exists_for_record(*criteria):
    return exists().where(
        MatchResult.consumption_record_id == ConsumptionRecord.id,
        *criteria,
    )


def _filter_records_by_student_no(q, student_no: str):
    """Filter consumption records by the student's normalized student number."""
    normalized_student_no = _normalize_text(student_no)
    return q.filter(ConsumptionRecord.student_no == normalized_student_no)


def _select_match_for_record(matches, status: str | None = None) -> MatchResult | None:
    match_items = list(matches or [])
    if status:
        status_matches = [
            match
            for match in match_items
            if (match.status.value if match.status else None) == status
        ]
        if status_matches:
            match_items = status_matches

    if not match_items:
        return None

    return max(match_items, key=lambda match: match.id or 0)


def _get_allowed_transaction_locations() -> list[str]:
    cfg = get_effective_config(current_app.config)
    return normalize_allowed_transaction_locations(cfg.get(CONSUMPTION_ALLOWED_LOCATIONS_KEY, []))


def _make_import_batch_id() -> str:
    from app.services.import_service import _resolve_import_timezone

    return datetime.now(_resolve_import_timezone()).strftime("%Y%m%d%H%M%S%f")[:-3]


def _delete_records_by_query(q) -> int:
    record_ids = [record_id for (record_id,) in q.with_entities(ConsumptionRecord.id).all()]
    if not record_ids:
        return 0

    MatchResult.query.filter(MatchResult.consumption_record_id.in_(record_ids)).delete(
        synchronize_session=False
    )
    ConsumptionRecord.query.filter(ConsumptionRecord.id.in_(record_ids)).delete(
        synchronize_session=False
    )
    db.session.commit()
    return len(record_ids)


@bp.route("/import-settings", methods=["GET"])
@role_required("admin")
def get_import_settings():
    return api_ok({
        "allowed_locations": _get_allowed_transaction_locations(),
    })


@bp.route("/import-settings", methods=["PUT"])
@role_required("admin")
def update_import_settings():
    data = request.get_json() or {}
    allowed_locations = normalize_allowed_transaction_locations(data.get("allowed_locations"))

    updates = {
        CONSUMPTION_ALLOWED_LOCATIONS_KEY: allowed_locations,
    }
    runtime_config_path = persist_runtime_overrides(current_app.config, updates)
    current_app.config.update(updates)
    current_app.config["LOCAL_RUNTIME_CONFIG_PATH"] = runtime_config_path

    return api_ok({
        "allowed_locations": allowed_locations,
        "runtime_config_path": runtime_config_path,
    })


@bp.route("/db-sync/status", methods=["GET"])
@role_required("admin")
def get_db_sync_status():
    from app.services.ztk_consumption_sync import ZtkConsumptionSyncService

    return api_ok(ZtkConsumptionSyncService().status())


@bp.route("/db-sync/trigger", methods=["POST"])
@role_required("admin")
def trigger_db_sync():
    from app.tasks.ztk_consumption import sync_ztk_consumption

    sync_ztk_consumption.delay(True)
    return api_ok({"message": "一卡通数据库同步任务已提交"})


ZTK_REQUIRED_KEYS = ("ZTK_DB_HOST", "ZTK_DB_NAME", "ZTK_DB_USER", "ZTK_DB_PASSWORD")


def _coerce_port(value) -> int:
    """Parse and range-check a port. Raises ValueError on invalid input."""
    try:
        port = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError("端口必须为整数") from exc
    if port < 1 or port > 65535:
        raise ValueError("端口必须在 1 到 65535 之间")
    return port


def _safe_port(value, default: int = 1433) -> int:
    """Best-effort port parse for read paths (e.g. hand-edited config); never raises."""
    try:
        return _coerce_port(value)
    except (TypeError, ValueError):
        return default


def _coerce_positive_int(value, label: str) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label}必须为整数") from exc
    if parsed < 1:
        raise ValueError(f"{label}必须大于等于 1")
    return parsed


def _safe_positive_int(value, default: int) -> int:
    try:
        return _coerce_positive_int(value, "数值")
    except (TypeError, ValueError):
        return default


def _ztk_db_sync_config_payload(cfg) -> dict:
    return {
        "host": cfg.get("ZTK_DB_HOST", ""),
        "port": _safe_port(cfg.get("ZTK_DB_PORT"), 1433),
        "database": cfg.get("ZTK_DB_NAME", ""),
        "user": cfg.get("ZTK_DB_USER", ""),
        "has_password": bool(_normalize_text(cfg.get("ZTK_DB_PASSWORD"))),
        "payment_books_table": cfg.get("ZTK_PAYMENT_BOOKS_TABLE", "dbo.view_ac_paymentbooks"),
        "sync_enabled": bool(cfg.get("ZTK_SYNC_ENABLED")),
        "sync_interval_minutes": _safe_positive_int(cfg.get("ZTK_SYNC_INTERVAL_MINUTES"), 5),
        "enabled_transaction_location_ids": get_enabled_transaction_location_ids(cfg),
        "configured": all(_normalize_text(cfg.get(key)) for key in ZTK_REQUIRED_KEYS),
    }


def _validate_ztk_table_name(value, label) -> str:
    """Validate a table identifier using the same rules as the sync service.

    Returns the normalized name; raises ValueError if empty or malformed.
    """
    from app.services.ztk_consumption_sync import _quote_table_name

    raw = _normalize_text(value)
    if not raw:
        raise ValueError(f"{label} 不能为空")
    try:
        _quote_table_name(raw, label)
    except RuntimeError as exc:
        raise ValueError(str(exc)) from exc
    return raw


@bp.route("/db-sync/config", methods=["GET"])
@role_required("admin")
def get_db_sync_config():
    cfg = get_effective_config(current_app.config)
    return api_ok(_ztk_db_sync_config_payload(cfg))


@bp.route("/db-sync/config", methods=["PUT"])
@role_required("admin")
def update_db_sync_config():
    data = request.get_json() or {}
    updates: dict[str, object] = {}

    try:
        if "host" in data:
            host = _normalize_text(data.get("host"))
            if not host:
                raise ValueError("主机地址不能为空")
            updates["ZTK_DB_HOST"] = host
        if "port" in data:
            updates["ZTK_DB_PORT"] = _coerce_port(data.get("port"))
        if "database" in data:
            database = _normalize_text(data.get("database"))
            if not database:
                raise ValueError("数据库名不能为空")
            updates["ZTK_DB_NAME"] = database
        if "user" in data:
            user = _normalize_text(data.get("user"))
            if not user:
                raise ValueError("用户名不能为空")
            updates["ZTK_DB_USER"] = user
        if "password" in data:
            password = str(data.get("password") or "")
            # Empty password means "keep the existing one"; only overwrite when provided.
            if password.strip():
                updates["ZTK_DB_PASSWORD"] = password
        if "payment_books_table" in data:
            updates["ZTK_PAYMENT_BOOKS_TABLE"] = _validate_ztk_table_name(
                data.get("payment_books_table"), "ZTK_PAYMENT_BOOKS_TABLE"
            )
        if "sync_enabled" in data:
            updates["ZTK_SYNC_ENABLED"] = bool(data.get("sync_enabled"))
        if "sync_interval_minutes" in data:
            updates["ZTK_SYNC_INTERVAL_MINUTES"] = _coerce_positive_int(
                data.get("sync_interval_minutes"), "同步间隔"
            )
        if "enabled_transaction_location_ids" in data:
            updates[ENABLED_TRANSACTION_LOCATION_IDS_KEY] = normalize_allowed_transaction_locations(
                data.get("enabled_transaction_location_ids")
            )
    except ValueError as e:
        return api_error(str(e))

    if not updates:
        return api_error("没有可更新的配置项")

    runtime_config_path = persist_runtime_overrides(current_app.config, updates)
    current_app.config.update(updates)
    current_app.config["LOCAL_RUNTIME_CONFIG_PATH"] = runtime_config_path

    cfg = get_effective_config(current_app.config)
    payload = _ztk_db_sync_config_payload(cfg)
    payload["runtime_config_path"] = runtime_config_path
    return api_ok(payload)


@bp.route("/db-sync/test", methods=["POST"])
@role_required("admin")
def test_db_sync():
    from app.services.ztk_consumption_sync import ZtkConsumptionSyncService

    data = request.get_json() or {}
    # Start from the effective (saved + env) config, then overlay any non-empty
    # fields from the request so admins can test edits before saving.
    test_config = get_effective_config(current_app.config)

    try:
        if _normalize_text(data.get("host")):
            test_config["ZTK_DB_HOST"] = _normalize_text(data.get("host"))
        if data.get("port") not in (None, ""):
            # Same range check as the save path so invalid ports are rejected
            # up front instead of producing an opaque pymssql error.
            test_config["ZTK_DB_PORT"] = _coerce_port(data.get("port"))
        if _normalize_text(data.get("database")):
            test_config["ZTK_DB_NAME"] = _normalize_text(data.get("database"))
        if _normalize_text(data.get("user")):
            test_config["ZTK_DB_USER"] = _normalize_text(data.get("user"))
        # Password: empty string means "use the saved one".
        if str(data.get("password") or "").strip():
            test_config["ZTK_DB_PASSWORD"] = str(data.get("password"))
        if _normalize_text(data.get("payment_books_table")):
            test_config["ZTK_PAYMENT_BOOKS_TABLE"] = _validate_ztk_table_name(
                data.get("payment_books_table"), "ZTK_PAYMENT_BOOKS_TABLE"
            )
    except ValueError as e:
        return api_error(str(e))

    try:
        result = ZtkConsumptionSyncService(config=test_config).test_connection()
    except Exception as exc:
        logger.error("ZTK db-sync test failed: %s", exc, exc_info=True)
        result = {
            "ok": False,
            "message": f"测试失败: {exc}",
            "latency_ms": 0.0,
            "server_version": None,
            "tables": {"payment_books": False},
        }
    return api_ok(result)


@bp.route("/import-template", methods=["GET"])
@role_required("admin")
def download_import_template():
    """Download Excel template for consumption record import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "消费记录导入模板"

    columns = [
        ("学号/消费卡号 *", "student_id"),
        ("学生姓名", "student_name"),
        ("消费时间 *", "transaction_time"),
        ("消费金额 *", "amount"),
        ("流水号 *", "transaction_id"),
        ("通道 *", "channel_id"),
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, (header, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    example_data = [
        "230501",
        "张三",
        "2026-03-31 12:05:30",
        12.50,
        "TX202603310001",
        "1",
    ]
    for col_idx, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border

    for idx, width in enumerate([18, 12, 22, 12, 22, 10], 1):
        ws.column_dimensions[chr(64 + idx)].width = width

    note_cell = ws.cell(
        row=4,
        column=1,
        value="说明：带 * 的字段为必填项。消费时间支持 YYYY-MM-DD HH:MM:SS、YYYY/MM/DD HH:MM:SS、YYYY-MM-DD HH:MM 等格式；通道需与视频通道 ID 一致。",
    )
    note_cell.font = Font(color="666666", italic=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="消费记录导入模板.xlsx",
    )


@bp.route("/import", methods=["POST"])
@role_required("admin")
def import_records():
    if "file" not in request.files:
        return api_error("请上传文件")

    file = request.files["file"]
    if not file.filename:
        return api_error("文件名不能为空")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("csv", "xls", "xlsx"):
        return api_error("仅支持 CSV、XLS、XLSX 格式")

    field_mapping = request.form.get("field_mapping")
    import json
    mapping = json.loads(field_mapping) if field_mapping else {}

    content = file.read()
    batch_id = _make_import_batch_id()

    try:
        svc = ConsumptionImportService()
        result = svc.import_file(
            content,
            ext,
            batch_id,
            mapping,
            allowed_locations=_get_allowed_transaction_locations(),
        )
    except Exception as e:
        logger.error(f"Import failed: {e}", exc_info=True)
        return api_error(f"导入失败：{str(e)}")

    # Trigger matching for imported records
    if result["imported"] > 0:
        from app.tasks.matching import run_matching_for_batch
        run_matching_for_batch.delay(batch_id)

    return api_ok(result)


@bp.route("/preview", methods=["POST"])
@role_required("admin")
def preview_import():
    """Preview first 10 rows of file before import."""
    if "file" not in request.files:
        return api_error("请上传文件")

    file = request.files["file"]
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("csv", "xls", "xlsx"):
        return api_error("仅支持 CSV、XLS、XLSX 格式")

    content = file.read()
    svc = ConsumptionImportService()
    try:
        preview = svc.preview(content, ext)
    except Exception as e:
        return api_error(f"文件解析失败：{str(e)}")
    return api_ok(preview)


@bp.route("/records", methods=["GET"])
@login_required
def list_records():
    q = apply_enabled_transaction_location_filter(
        ConsumptionRecord.query
    ).order_by(ConsumptionRecord.transaction_time.desc())
    if student_no := request.args.get("student_no"):
        q = _filter_records_by_student_no(q, student_no)
    elif student_id := request.args.get("student_id"):
        q = q.filter(ConsumptionRecord.student_id == student_id)
    if student_query := (request.args.get("student") or request.args.get("student_query")):
        student_query = student_query.strip()
        if student_query:
            like = f"%{student_query}%"
            q = q.filter(or_(
                ConsumptionRecord.student_no.like(like),
                ConsumptionRecord.student_name.like(like),
            ))
    if date_str := request.args.get("date"):
        try:
            d = date.fromisoformat(date_str)
            q = q.filter(db.func.date(ConsumptionRecord.transaction_time) == d)
        except ValueError:
            return api_error("日期格式无效")
    else:
        start_date_str = request.args.get("date_from") or request.args.get("start_date")
        end_date_str = request.args.get("date_to") or request.args.get("end_date")
        try:
            start_date = date.fromisoformat(start_date_str) if start_date_str else None
            end_date = date.fromisoformat(end_date_str) if end_date_str else None
        except ValueError:
            return api_error("日期格式无效")
        if start_date and end_date and start_date > end_date:
            return api_error("开始日期不能晚于结束日期")
        if start_date:
            q = q.filter(db.func.date(ConsumptionRecord.transaction_time) >= start_date)
        if end_date:
            q = q.filter(db.func.date(ConsumptionRecord.transaction_time) <= end_date)
    if batch := (request.args.get("batch") or request.args.get("import_batch")):
        q = q.filter(ConsumptionRecord.import_batch == batch)
    if channel_id := (request.args.get("channel_id") or request.args.get("channel")):
        channel_id = channel_id.strip()
        if channel_id:
            q = q.filter(ConsumptionRecord.channel_id.like(f"%{channel_id}%"))
    if transaction_id := request.args.get("transaction_id"):
        transaction_id = transaction_id.strip()
        if transaction_id:
            q = q.filter(ConsumptionRecord.transaction_id.like(f"%{transaction_id}%"))

    items, total, page, page_size = paginate(q)
    return api_ok(paginated_response([r.to_dict() for r in items], total, page, page_size))


@bp.route("/records/batches", methods=["GET"])
@login_required
def list_record_batches():
    q = db.session.query(
        ConsumptionRecord.import_batch.label("batch_id"),
        db.func.count(ConsumptionRecord.id).label("record_count"),
        db.func.min(ConsumptionRecord.transaction_time).label("first_transaction_time"),
        db.func.max(ConsumptionRecord.transaction_time).label("last_transaction_time"),
        db.func.max(ConsumptionRecord.created_at).label("created_at"),
        db.func.sum(ConsumptionRecord.amount).label("total_amount"),
    )
    q = apply_enabled_transaction_location_filter(q).filter(
        ConsumptionRecord.import_batch.isnot(None),
        ConsumptionRecord.import_batch != "",
    )

    if batch := request.args.get("batch"):
        q = q.filter(ConsumptionRecord.import_batch.like(f"%{batch}%"))

    rows = q.group_by(ConsumptionRecord.import_batch).order_by(
        db.func.max(ConsumptionRecord.created_at).desc()
    ).limit(200).all()

    return api_ok({
        "items": [
            {
                "batch_id": row.batch_id,
                "record_count": row.record_count,
                "first_transaction_time": row.first_transaction_time.isoformat() if row.first_transaction_time else None,
                "last_transaction_time": row.last_transaction_time.isoformat() if row.last_transaction_time else None,
                "created_at": row.created_at.isoformat() if row.created_at else None,
                "total_amount": float(row.total_amount) if row.total_amount is not None else 0.0,
            }
            for row in rows
        ]
    })


@bp.route("/records/<int:record_id>", methods=["DELETE"])
@role_required("admin")
def delete_record(record_id):
    record = ConsumptionRecord.query.get_or_404(record_id)
    batch_id = record.import_batch
    transaction_id = record.transaction_id

    MatchResult.query.filter_by(consumption_record_id=record.id).delete()
    db.session.delete(record)
    db.session.commit()

    return api_ok({
        "id": record_id,
        "batch_id": batch_id,
        "transaction_id": transaction_id,
        "deleted": 1,
    })


@bp.route("/records/batches/<path:batch_id>", methods=["DELETE"])
@role_required("admin")
def delete_record_batch(batch_id):
    deleted = _delete_records_by_query(
        ConsumptionRecord.query.filter(ConsumptionRecord.import_batch == batch_id)
    )
    return api_ok({
        "batch_id": batch_id,
        "deleted": deleted,
    })


@bp.route("/matches", methods=["GET"])
@login_required
def list_matches():
    q = ConsumptionRecord.query.options(
        selectinload(ConsumptionRecord.match_result).selectinload(MatchResult.image),
        joinedload(ConsumptionRecord.student),
    )
    q = apply_enabled_transaction_location_filter(q).filter(
        ConsumptionRecord.amount < 0,
    ).order_by(ConsumptionRecord.transaction_time.desc())

    if date_str := request.args.get("date"):
        try:
            d = date.fromisoformat(date_str)
            q = q.filter(db.func.date(ConsumptionRecord.transaction_time) == d)
        except ValueError:
            return api_error("日期格式无效")
    if student_no := request.args.get("student_no"):
        q = _filter_records_by_student_no(q, student_no)
    elif student_id := request.args.get("student_id"):
        q = q.filter(ConsumptionRecord.student_id == student_id)
    if status := request.args.get("status"):
        if status == MatchStatusEnum.unmatched_record.value:
            q = q.filter(or_(
                ~_match_exists_for_record(),
                _match_exists_for_record(MatchResult.status == MatchStatusEnum.unmatched_record),
            ))
        else:
            q = q.filter(_match_exists_for_record(MatchResult.status == status))

    items, total, page, page_size = paginate(q)
    result = []
    for record in items:
        match = _select_match_for_record(record.match_result, status)
        if match:
            d = _build_match_payload(match, record)
        else:
            d = {
                "id": record.id,
                "consumption_record_id": record.id,
                "image_id": None,
                "student_id": record.student_id,
                "status": MatchStatusEnum.unmatched_record.value,
                "time_diff_seconds": None,
                "price_diff": None,
                "image_price_total": None,
                "is_manual": False,
                "match_date": record.transaction_time.date().isoformat() if record.transaction_time else None,
                "created_at": record.created_at.isoformat() if record.created_at else None,
            }

        d["consumption_record"] = record.to_dict()
        if record.student:
            d["student"] = record.student.to_dict()
        if match and match.image:
            image = match.image.to_dict()
            recs = DishRecognition.query.filter_by(image_id=match.image.id).all()
            image["recognitions"] = [r.to_dict() for r in recs]
            d["image"] = image
        result.append(d)
    return api_ok(paginated_response(result, total, page, page_size))


@bp.route("/matches/unmatched-images", methods=["GET"])
@login_required
def list_unmatched_images():
    q = MatchResult.query.options(
        joinedload(MatchResult.image),
    ).filter(
        MatchResult.status == MatchStatusEnum.unmatched_image,
    ).order_by(MatchResult.created_at.desc())

    if date_str := request.args.get("date"):
        try:
            d = date.fromisoformat(date_str)
            q = q.filter(MatchResult.match_date == d)
        except ValueError:
            return api_error("日期格式无效")

    items, total, page, page_size = paginate(q)
    result = []
    for match in items:
        d = _build_match_payload(match)
        if match.image:
            image = match.image.to_dict()
            recs = DishRecognition.query.filter_by(image_id=match.image.id).all()
            image["recognitions"] = [r.to_dict() for r in recs]
            d["image"] = image
        result.append(d)
    return api_ok(paginated_response(result, total, page, page_size))


@bp.route("/matches/<int:match_id>", methods=["GET"])
@login_required
def get_match(match_id):
    match = MatchResult.query.options(
        joinedload(MatchResult.consumption_record).joinedload(ConsumptionRecord.student),
        joinedload(MatchResult.image),
        joinedload(MatchResult.student),
    ).filter(MatchResult.id == match_id).first_or_404()
    return api_ok(_build_match_detail_payload(match))


@bp.route("/matches/<int:match_id>/confirm", methods=["PUT"])
@role_required("admin")
def confirm_match(match_id):
    m = MatchResult.query.get_or_404(match_id)
    data = request.get_json() or {}

    if data.get("image_id"):
        m.image_id = data["image_id"]

    m.status = MatchStatusEnum.confirmed
    m.is_manual = True
    m.confirmed_by = request.current_user.id
    from datetime import datetime, timezone
    m.confirmed_at = datetime.now(timezone.utc)
    db.session.commit()

    # Recompute nutrition log
    if m.student_id and m.match_date:
        from app.tasks.nutrition import compute_nutrition_log
        compute_nutrition_log.delay(m.student_id, m.match_date.isoformat())

    return api_ok(m.to_dict())


@bp.route("/matches/rematch", methods=["POST"])
@role_required("admin")
def rematch():
    """Batch re-trigger matching for a date."""
    data = request.get_json() or {}
    date_str = data.get("date", date.today().isoformat())
    from app.tasks.matching import run_matching_for_date
    run_matching_for_date.delay(date_str)
    return api_ok({"message": f"已触发 {date_str} 的重新匹配任务"})
