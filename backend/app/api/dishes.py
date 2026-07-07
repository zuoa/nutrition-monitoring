import io
import logging
import os
import re
import tempfile
import unicodedata
import uuid
import zipfile
from datetime import datetime
from pathlib import Path
from flask import Blueprint, request, current_app, send_file
from app import db
from app.models import Dish, CategoryEnum, DishSampleImage, EmbeddingStatusEnum, TaskLog
from app.utils.jwt_utils import login_required, role_required, api_ok, api_error
from app.utils.pagination import paginate, paginated_response
from app.services.dish_analyzer import DishAnalyzerService
from app.services.embedding_jobs import can_trigger_local_embedding_rebuild, trigger_local_embedding_rebuild
from app.services.qwen_vl import QwenVLService
from app.services.structured_description import compose_structured_description, empty_structured_description
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

bp = Blueprint("dishes", __name__)
logger = logging.getLogger(__name__)

ALLOWED_ROLES_WRITE = ("admin", "canteen_manager")
MAX_DISH_SAMPLE_IMAGES = 12
ALLOWED_SAMPLE_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}


@bp.route("/", methods=["GET"])
@login_required
def list_dishes():
    q = Dish.query
    # Filters
    if request.args.get("active_only") != "false":
        q = q.filter(Dish.is_active)
    if category := request.args.get("category"):
        q = q.filter(Dish.category == category)
    if search := request.args.get("search"):
        q = q.filter(Dish.name.ilike(f"%{search}%"))
    has_sample_images = request.args.get("has_sample_images")
    if has_sample_images in {"true", "false"}:
        has_active_sample_images = Dish.sample_images.any(DishSampleImage.is_active.is_(True))
        q = q.filter(has_active_sample_images if has_sample_images == "true" else ~has_active_sample_images)
    q = q.order_by(Dish.category, Dish.name)

    items, total, page, page_size = paginate(q)
    return api_ok(paginated_response([d.to_dict() for d in items], total, page, page_size))


@bp.route("/<int:dish_id>", methods=["GET"])
@login_required
def get_dish(dish_id):
    dish = Dish.query.get_or_404(dish_id)
    return api_ok(dish.to_dict())


@bp.route("/", methods=["POST"])
@role_required(*ALLOWED_ROLES_WRITE)
def create_dish():
    data = request.get_json() or {}
    errors = _validate_dish(data)
    if errors:
        return api_error("; ".join(errors))

    name = data["name"].strip()
    if Dish.query.filter(Dish.name.ilike(name)).first():
        return api_error(f"菜品「{name}」已存在")

    dish = Dish(
        name=name,
        description=data.get("description"),
        ingredients=data.get("ingredients"),
        image_url=data.get("image_url"),
        price=data["price"],
        category=data["category"],
        weight=data.get("weight", 100),
        calories=data.get("calories"),
        protein=data.get("protein"),
        fat=data.get("fat"),
        carbohydrate=data.get("carbohydrate"),
        sodium=data.get("sodium"),
        fiber=data.get("fiber"),
    )
    db.session.add(dish)
    db.session.commit()
    return api_ok(dish.to_dict()), 201


@bp.route("/<int:dish_id>", methods=["PUT"])
@role_required(*ALLOWED_ROLES_WRITE)
def update_dish(dish_id):
    dish = Dish.query.get_or_404(dish_id)
    data = request.get_json() or {}

    if "name" in data:
        name = data["name"].strip()
        existing = Dish.query.filter(Dish.name.ilike(name), Dish.id != dish_id).first()
        if existing:
            return api_error(f"菜品「{name}」已存在")
        dish.name = name

    for field in ["description", "ingredients", "image_url", "price", "category", "weight",
                  "calories", "protein", "fat", "carbohydrate", "sodium", "fiber", "is_active"]:
        if field in data:
            setattr(dish, field, data[field])

    db.session.commit()
    return api_ok(dish.to_dict())


@bp.route("/<int:dish_id>", methods=["DELETE"])
@role_required(*ALLOWED_ROLES_WRITE)
def delete_dish(dish_id):
    dish = Dish.query.get_or_404(dish_id)
    dish.is_active = False  # soft delete
    db.session.commit()
    return api_ok({"id": dish_id})


@bp.route("/categories", methods=["GET"])
@login_required
def list_categories():
    return api_ok([c.value for c in CategoryEnum])


def _validate_dish(data):
    errors = []
    if not data.get("name", "").strip():
        errors.append("菜品名称不能为空")
    if data.get("price") is None:
        errors.append("价格不能为空")
    elif float(data["price"]) < 0:
        errors.append("价格不能为负数")
    if not data.get("category"):
        errors.append("分类不能为空")
    elif data["category"] not in [c.value for c in CategoryEnum]:
        errors.append(f"分类无效，可选：{[c.value for c in CategoryEnum]}")
    return errors


def _validate_sample_image_file(file_storage):
    filename = (file_storage.filename or "").strip()
    if not filename:
        return "文件名无效"

    ext = os.path.splitext(filename)[1].lower()
    if ext not in ALLOWED_SAMPLE_IMAGE_EXTENSIONS:
        return f"不支持的图片格式，请上传 {', '.join(sorted(ALLOWED_SAMPLE_IMAGE_EXTENSIONS))} 格式"

    try:
        current_pos = file_storage.stream.tell()
        file_storage.stream.seek(0, os.SEEK_END)
        size = file_storage.stream.tell()
        file_storage.stream.seek(current_pos)
    except (AttributeError, OSError):
        size = 0

    max_size = current_app.config.get("MAX_IMAGE_SIZE", 5 * 1024 * 1024)
    if size and size > max_size:
        return f"图片大小不能超过 {max_size // (1024 * 1024)}MB"

    return None


def _save_sample_image_file(dish_id: int, file_storage, *, sort_order: int, is_cover: bool) -> DishSampleImage:
    ext = os.path.splitext(file_storage.filename or "")[1].lower()
    image_root = current_app.config.get("IMAGE_STORAGE_PATH", "/data/images")
    dest_dir = os.path.join(image_root, "dish_samples", str(dish_id))
    os.makedirs(dest_dir, exist_ok=True)

    stored_name = f"{uuid.uuid4().hex}{ext}"
    dest_path = os.path.join(dest_dir, stored_name)
    file_storage.save(dest_path)

    return DishSampleImage(
        dish_id=dish_id,
        image_path=dest_path,
        original_filename=file_storage.filename,
        sort_order=sort_order,
        is_cover=is_cover,
        embedding_status=EmbeddingStatusEnum.pending,
    )


def _reset_sample_image_embedding_state(image: DishSampleImage):
    image.embedding_status = EmbeddingStatusEnum.pending
    image.embedding_model = None
    image.embedding_version = None
    image.embedding_input_hash = None
    image.embedding_vector = None
    image.embedding_updated_at = None
    image.error_message = None


def _replace_sample_image_file(image: DishSampleImage, file_storage):
    ext = os.path.splitext(file_storage.filename or "")[1].lower()
    image_root = current_app.config.get("IMAGE_STORAGE_PATH", "/data/images")
    dest_dir = os.path.join(image_root, "dish_samples", str(image.dish_id))
    os.makedirs(dest_dir, exist_ok=True)

    stored_name = f"{uuid.uuid4().hex}{ext}"
    dest_path = os.path.join(dest_dir, stored_name)
    file_storage.save(dest_path)

    old_path = image.image_path
    image.image_path = dest_path
    image.original_filename = file_storage.filename
    _reset_sample_image_embedding_state(image)

    if old_path and old_path != dest_path and os.path.exists(old_path):
        try:
            os.unlink(old_path)
        except OSError as e:
            logger.warning("Failed to delete previous sample image file %s: %s", old_path, e)


def _delete_sample_image_file(image: DishSampleImage):
    image_path = image.image_path
    if image_path and os.path.exists(image_path):
        try:
            os.unlink(image_path)
        except OSError as e:
            logger.warning("Failed to delete sample image file %s: %s", image_path, e)


def _ensure_cover_image(dish_id: int):
    has_cover = db.session.query(DishSampleImage.id).filter(
        DishSampleImage.dish_id == dish_id,
        DishSampleImage.is_active.is_(True),
        DishSampleImage.is_cover.is_(True),
    ).first()
    if has_cover:
        return

    next_image = DishSampleImage.query.filter_by(
        dish_id=dish_id,
        is_active=True,
    ).order_by(DishSampleImage.sort_order.asc(), DishSampleImage.id.asc()).first()
    if next_image:
        next_image.is_cover = True


@bp.route("/<int:dish_id>/images", methods=["POST"])
@role_required(*ALLOWED_ROLES_WRITE)
def upload_dish_images(dish_id):
    dish = Dish.query.get_or_404(dish_id)
    files = request.files.getlist("images")
    if not files and "image" in request.files:
        files = [request.files["image"]]

    files = [f for f in files if (f.filename or "").strip()]
    if not files:
        return api_error("请至少上传一张图片")

    active_count = DishSampleImage.query.filter_by(dish_id=dish.id, is_active=True).count()
    if active_count + len(files) > MAX_DISH_SAMPLE_IMAGES:
        return api_error(f"每个菜品最多上传 {MAX_DISH_SAMPLE_IMAGES} 张样图")

    for file in files:
        error = _validate_sample_image_file(file)
        if error:
            return api_error(f"{file.filename}: {error}")

    created_images = []
    try:
        current_max_sort = db.session.query(db.func.max(DishSampleImage.sort_order)).filter(
            DishSampleImage.dish_id == dish.id,
            DishSampleImage.is_active.is_(True),
        ).scalar() or 0
        has_cover = db.session.query(DishSampleImage.id).filter(
            DishSampleImage.dish_id == dish.id,
            DishSampleImage.is_cover.is_(True),
            DishSampleImage.is_active.is_(True),
        ).first()
        for file in files:
            current_max_sort += 1
            image = _save_sample_image_file(
                dish.id,
                file,
                sort_order=current_max_sort,
                is_cover=not bool(has_cover),
            )
            db.session.add(image)
            created_images.append(image)
            has_cover = True
        db.session.commit()

        try:
            trigger_local_embedding_rebuild(current_app.config, reason="dish sample upload")
        except Exception as e:
            logger.warning("Failed to trigger local embedding rebuild after upload: %s", e)
    except Exception as e:
        db.session.rollback()
        for image in created_images:
            _delete_sample_image_file(image)
        logger.error("Failed to upload dish sample images for dish %s: %s", dish.id, e)
        return api_error(f"上传样图失败: {str(e)}"), 500

    return api_ok({
        "dish_id": dish.id,
        "images": [img.to_dict() for img in created_images],
        "sample_image_count": DishSampleImage.query.filter_by(dish_id=dish.id, is_active=True).count(),
    }), 201


@bp.route("/images/<int:image_id>", methods=["DELETE"])
@role_required(*ALLOWED_ROLES_WRITE)
def delete_dish_image(image_id):
    image = DishSampleImage.query.get_or_404(image_id)
    dish_id = image.dish_id
    _delete_sample_image_file(image)
    db.session.delete(image)
    db.session.flush()
    _ensure_cover_image(dish_id)
    db.session.commit()
    try:
        trigger_local_embedding_rebuild(current_app.config, reason="dish sample delete")
    except Exception as e:
        logger.warning("Failed to trigger local embedding rebuild after delete: %s", e)
    return api_ok({"id": image_id, "dish_id": dish_id})


@bp.route("/images/<int:image_id>", methods=["PUT"])
@role_required(*ALLOWED_ROLES_WRITE)
def update_dish_image(image_id):
    image = DishSampleImage.query.get_or_404(image_id)
    file = request.files.get("image")
    if file is None or not (file.filename or "").strip():
        return api_error("请上传图片")

    error = _validate_sample_image_file(file)
    if error:
        return api_error(error)

    try:
        _replace_sample_image_file(image, file)
        db.session.commit()

        try:
            trigger_local_embedding_rebuild(current_app.config, reason="dish sample replace")
        except Exception as e:
            logger.warning("Failed to trigger local embedding rebuild after replace: %s", e)
    except Exception as e:
        db.session.rollback()
        logger.error("Failed to replace dish sample image %s: %s", image_id, e, exc_info=True)
        return api_error(f"更新样图失败: {str(e)}"), 500

    return api_ok({
        "dish_id": image.dish_id,
        "image": image.to_dict(),
    })


@bp.route("/rebuild-sample-embeddings", methods=["POST"])
@role_required(*ALLOWED_ROLES_WRITE)
def rebuild_dish_sample_embeddings():
    allowed, skip_reason = can_trigger_local_embedding_rebuild(current_app.config)
    if not allowed:
        return api_error(f"当前配置不支持重建本地样图 embedding: {skip_reason}")

    try:
        from app.tasks.embeddings import rebuild_sample_embeddings
        rebuild_sample_embeddings.delay()
        return api_ok({"message": "样图 embedding 重建任务已提交"})
    except Exception as e:
        logger.error("Failed to submit embedding rebuild task: %s", e, exc_info=True)
        return api_error(f"提交重建任务失败: {str(e)}"), 500


@bp.route("/<int:dish_id>/analyze-nutrition", methods=["POST"])
@role_required(*ALLOWED_ROLES_WRITE)
def analyze_dish_nutrition(dish_id):
    """Analyze dish nutrition using AI and update dish record."""
    dish = Dish.query.get_or_404(dish_id)
    data = request.get_json() or {}
    weight = int(data.get("weight", 100))

    if weight <= 0 or weight > 10000:
        return api_error("重量必须在 1-10000g 之间")

    # Get config from app
    config = current_app.config
    api_key = config.get("OPENAI_API_KEY", "")

    if not api_key:
        return api_error("营养分析服务未配置 (OPENAI_API_KEY)"), 503

    try:
        analyzer = DishAnalyzerService(config)
        result = analyzer.analyze_nutrition(dish.name, weight)

        # Update dish with analyzed nutrition data and description
        dish.weight = weight
        dish.calories = result.get("calories")
        dish.protein = result.get("protein")
        dish.fat = result.get("fat")
        dish.carbohydrate = result.get("carbohydrate")
        dish.sodium = result.get("sodium")
        dish.fiber = result.get("fiber")
        composed_description = compose_structured_description(
            result.get("description", ""),
            result.get("structured_description"),
        )
        if composed_description:
            dish.description = composed_description

        db.session.commit()

        return api_ok({
            "dish": dish.to_dict(),
            "weight": weight,
            "structured_description": result.get("structured_description", {}),
            "analysis_notes": result.get("notes", ""),
        })
    except Exception as e:
        logger.error(f"Failed to analyze dish nutrition: {e}")
        return api_error(f"营养分析失败: {str(e)}"), 500


@bp.route("/import-template", methods=["GET"])
@login_required
def download_import_template():
    """Download Excel template for dish import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "菜品导入模板"

    # Define columns with Chinese headers
    columns = [
        ("菜品名称 *", "name"),
        ("分类 *", "category"),
        ("单价(元) *", "price"),
        ("份量(g)", "weight"),
        ("视觉描述", "description"),
        ("配菜描述", "ingredients"),
        ("热量(kcal)", "calories"),
        ("蛋白质(g)", "protein"),
        ("脂肪(g)", "fat"),
        ("碳水化合物(g)", "carbohydrate"),
        ("钠(mg)", "sodium"),
        ("膳食纤维(g)", "fiber"),
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, (header, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Add example row with sample data
    example_data = [
        "红烧肉", "荤菜", 12.00, 150,
        "深红色酱汁包裹的五花肉块，肥瘦相间，表面油亮",
        "五花肉、土豆、冰糖、酱油",
        450, 25, 35, 8, 800, 1.5,
    ]
    for col_idx, value in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border

    # Set column widths
    col_widths = [15, 8, 10, 8, 40, 30, 12, 12, 10, 14, 10, 12]
    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + idx) if idx <= 26 else "A" + chr(64 + idx - 26)].width = width

    # Add a note row
    note_cell = ws.cell(row=4, column=1, value="说明：带 * 的字段为必填项。分类可选值：主食、荤菜、素菜、汤、其他")
    note_cell.font = Font(color="666666", italic=True)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="菜品导入模板.xlsx",
    )


# Excel column mapping (Chinese header -> field name), shared by Excel + ZIP import.
_DISH_IMPORT_COLUMN_MAPPING = {
    "菜品名称": "name", "菜品名称 *": "name",
    "分类": "category", "分类 *": "category",
    "单价(元)": "price", "单价(元) *": "price", "单价": "price",
    "份量(g)": "weight", "份量": "weight", "重量(g)": "weight",
    "视觉描述": "description",
    "配菜描述": "ingredients",
    "热量(kcal)": "calories", "热量": "calories",
    "蛋白质(g)": "protein", "蛋白质": "protein",
    "脂肪(g)": "fat", "脂肪": "fat",
    "碳水化合物(g)": "carbohydrate", "碳水化合物": "carbohydrate",
    "钠(mg)": "sodium", "钠": "sodium",
    "膳食纤维(g)": "fiber", "膳食纤维": "fiber",
}


def _normalize_dish_name(name):
    """Normalize a dish/folder name for matching: NFKC + collapse whitespace + casefold."""
    if not name:
        return ""
    s = unicodedata.normalize("NFKC", str(name)).strip()
    s = re.sub(r"\s+", " ", s)
    return s.casefold()


def _decode_zip_name(zi: zipfile.ZipInfo) -> str:
    """Return the human-readable filename for a ZIP entry.

    Python's zipfile decodes names as UTF-8 when flag bit 11 (0x800) is set,
    otherwise as CP437. macOS Finder writes UTF-8 filenames WITHOUT setting
    that bit, so Chinese names arrive as CP437 mojibake
    (e.g. 板栗炒鸡 -> 'µ¥┐µáùτéÆΘ╕í') and no longer match dish names read
    from Excel. Recover the original bytes (CP437 round-trips losslessly for
    all 256 byte values) and re-decode as UTF-8, falling back to GBK/GB18030
    for Windows-produced Chinese archives. Returns the original name if no
    re-decode succeeds."""
    name = zi.filename
    if zi.flag_bits & 0x800:
        return name  # already correctly UTF-8 decoded
    try:
        raw = name.encode("cp437")
    except (UnicodeEncodeError, LookupError):
        return name
    for enc in ("utf-8", "gbk", "gb18030"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return name


def _parse_optional_number(val, default=None):
    try:
        return float(val) if val and str(val).strip() else default
    except (ValueError, TypeError):
        return default


def _upsert_dishes_from_dataframe(df):
    """Validate + upsert dishes from a cleaned (fillna'd) Excel DataFrame.

    Stages changes in the session (does NOT commit). Returns
    (created_names, updated_names, errors, name_normalized -> Dish map).
    Callers must flush()/commit() to populate ids on newly created dishes.
    """
    df.columns = [_DISH_IMPORT_COLUMN_MAPPING.get(str(c).strip(), str(c).strip()) for c in df.columns]
    valid_categories = [c.value for c in CategoryEnum]
    errors = []
    created = []
    updated = []
    dish_map = {}

    for idx, row in df.iterrows():
        row_num = idx + 2  # Excel row number (1-indexed + header)

        name = row.get("name", "").strip()
        category = row.get("category", "").strip()
        price_str = str(row.get("price", "")).strip()

        if not name:
            errors.append(f"第{row_num}行: 菜品名称不能为空")
            continue
        if not category:
            errors.append(f"第{row_num}行: 分类不能为空")
            continue
        if category not in valid_categories:
            errors.append(f"第{row_num}行: 分类「{category}」无效，可选: {valid_categories}")
            continue
        if not price_str:
            errors.append(f"第{row_num}行: 单价不能为空")
            continue

        try:
            price = float(price_str)
            if price < 0:
                raise ValueError()
        except ValueError:
            errors.append(f"第{row_num}行: 单价格式无效")
            continue

        weight = _parse_optional_number(row.get("weight"), 100)
        calories = _parse_optional_number(row.get("calories"))
        protein = _parse_optional_number(row.get("protein"))
        fat = _parse_optional_number(row.get("fat"))
        carbohydrate = _parse_optional_number(row.get("carbohydrate"))
        sodium = _parse_optional_number(row.get("sodium"))
        fiber = _parse_optional_number(row.get("fiber"))

        existing = Dish.query.filter(Dish.name.ilike(name)).first()

        if existing:
            # Update existing dish
            existing.price = price
            existing.category = category
            existing.weight = weight
            if row.get("description"):
                existing.description = row.get("description")
            if row.get("ingredients"):
                existing.ingredients = row.get("ingredients")
            if calories is not None:
                existing.calories = calories
            if protein is not None:
                existing.protein = protein
            if fat is not None:
                existing.fat = fat
            if carbohydrate is not None:
                existing.carbohydrate = carbohydrate
            if sodium is not None:
                existing.sodium = sodium
            if fiber is not None:
                existing.fiber = fiber
            existing.is_active = True
            updated.append(name)
            dish_map[_normalize_dish_name(name)] = existing
        else:
            # Create new dish
            dish = Dish(
                name=name,
                description=row.get("description") or None,
                ingredients=row.get("ingredients") or None,
                price=price,
                category=category,
                weight=weight,
                calories=calories,
                protein=protein,
                fat=fat,
                carbohydrate=carbohydrate,
                sodium=sodium,
                fiber=fiber,
            )
            db.session.add(dish)
            created.append(name)
            dish_map[_normalize_dish_name(name)] = dish

    return created, updated, errors, dish_map


@bp.route("/import", methods=["POST"])
@role_required(*ALLOWED_ROLES_WRITE)
def import_dishes():
    """Import dishes from Excel file."""
    if "file" not in request.files:
        return api_error("未上传文件")

    file = request.files["file"]
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        return api_error("请上传 Excel 文件 (.xlsx 或 .xls)")

    try:
        df = pd.read_excel(file, sheet_name=0, dtype=str)
        df = df.fillna("").map(lambda x: str(x).strip() if x else "")
    except Exception as e:
        logger.error(f"Failed to parse Excel: {e}")
        return api_error(f"解析 Excel 失败: {str(e)}")

    created, updated, errors, _ = _upsert_dishes_from_dataframe(df)

    if errors and not created and not updated:
        return api_error("导入失败:\n" + "\n".join(errors[:20]))

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        logger.error(f"Failed to commit import: {e}")
        return api_error(f"保存失败: {str(e)}")

    result = {
        "created_count": len(created),
        "updated_count": len(updated),
        "created": created[:20],
        "updated": updated[:20],
    }
    if errors:
        result["warnings"] = errors[:10]

    return api_ok(result)


class _BytesFileStorage:
    """Minimal werkzeug.FileStorage stand-in wrapping in-memory bytes, so the
    existing _validate_sample_image_file / _save_sample_image_file helpers work
    unchanged for images read out of a ZIP archive."""

    def __init__(self, filename, data):
        self.filename = filename
        self._buf = io.BytesIO(data)

    @property
    def stream(self):
        return self._buf

    def save(self, dest_path):
        with open(dest_path, "wb") as f:
            f.write(self._buf.getvalue())


def _validate_image_magic_bytes(data):
    """Return None if data looks like a real image, else an error message.

    Degrades gracefully: if libmagic is unavailable (common on dev hosts
    without libmagic installed), trust the file extension and return None.
    """
    if not data:
        return "图片内容为空"
    try:
        import magic  # python-magic; lazy import (libmagic may be absent)
        mime = magic.from_buffer(data, mime=True)
    except Exception:
        return None
    if mime in {"image/jpeg", "image/png", "image/webp"}:
        return None
    return f"文件内容不是有效图片 ({mime})"


@bp.route("/import-zip", methods=["POST"])
@role_required(*ALLOWED_ROLES_WRITE)
def import_dishes_zip():
    """Async ZIP dish import for vectorization.

    The request only streams the upload to the shared image volume and enqueues
    a Celery task; the heavy work (zip validation, Excel parse, image save,
    embedding rebuild) runs in the worker. Returns a TaskLog id the client polls
    via GET /v1/analysis/tasks/<id>.

    ZIP layout: one Excel (.xlsx/.xls) + top-level folders named after each
    dish, each holding jpg/png sample images.
    """
    if "file" not in request.files:
        return api_error("未上传文件")

    file = request.files["file"]
    filename = (file.filename or "").strip()
    if not filename or not filename.lower().endswith(".zip"):
        return api_error("请上传 ZIP 文件 (.zip)")

    # Reject a duplicate upload while one is already pending/running.
    active = TaskLog.query.filter(
        TaskLog.task_type == "dish_zip_import",
        TaskLog.status.in_(("pending", "running")),
    ).order_by(TaskLog.started_at.desc()).first()
    if active:
        return api_ok({
            "task_id": active.id,
            "task": active.to_dict(),
            "message": "已有 ZIP 导入任务正在执行",
        })

    max_zip_size = current_app.config.get("MAX_IMPORT_ZIP_SIZE", 2 * 1024 * 1024 * 1024)
    image_root = current_app.config.get("IMAGE_STORAGE_PATH", "/data/images")
    upload_dir = os.path.join(image_root, "import_uploads")
    os.makedirs(upload_dir, exist_ok=True)
    saved_path = os.path.join(upload_dir, f"{uuid.uuid4().hex}.zip")

    try:
        # Stream to the shared volume in chunks (never hold a multi-GB archive in RAM).
        total = 0
        too_big = False
        with open(saved_path, "wb") as out:
            while True:
                chunk = file.stream.read(1024 * 1024)
                if not chunk:
                    break
                total += len(chunk)
                if total > max_zip_size:
                    too_big = True
                    break
                out.write(chunk)

        if too_big:
            return api_error(f"ZIP 文件过大，不能超过 {max_zip_size // (1024 * 1024)}MB")
        if total == 0:
            return api_error("上传文件为空")

        # Quick central-directory check; deep validation (testzip, Excel scan,
        # path checks) runs in the task.
        try:
            with zipfile.ZipFile(saved_path):
                pass
        except zipfile.BadZipFile:
            return api_error("无法读取 ZIP 文件，请确认是有效的 .zip")

        from app.tasks.dishes import create_zip_import_task_log, import_dishes_zip_task

        task_log = create_zip_import_task_log()
        try:
            celery_task = import_dishes_zip_task.delay(task_log.id, saved_path)
            saved_path = None  # ownership transferred to the task (it deletes it)
            task_log.meta = {**(task_log.meta or {}), "celery_task_id": celery_task.id}
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            task_log = TaskLog.query.get(task_log.id)
            if task_log:
                task_log.status = "failed"
                task_log.error_count = 1
                task_log.error_message = f"提交导入任务失败: {str(e)}"
                task_log.finished_at = datetime.utcnow()
                db.session.commit()
            logger.error("Failed to submit zip import task: %s", e, exc_info=True)
            return api_error(f"提交导入任务失败: {str(e)}"), 500

        return api_ok({
            "task_id": task_log.id,
            "task": task_log.to_dict(),
            "message": "ZIP 导入任务已提交",
        })
    finally:
        # Clean up only if the task never took ownership of the file.
        if saved_path and os.path.exists(saved_path):
            try:
                os.unlink(saved_path)
            except OSError:
                pass


def _execute_zip_import(zip_path, task_log):
    """Parse the zip at zip_path, upsert dishes from its Excel, and import
    per-dish sample images. Writes phase progress to task_log.meta (reassigned,
    then committed) so the status endpoint can report it. Raises ValueError on
    validation failure (the Celery wrapper marks the task failed). Returns a
    result dict."""
    max_entries = current_app.config.get("MAX_ZIP_ENTRIES", 5000)
    max_extracted = current_app.config.get("MAX_ZIP_EXTRACTED_SIZE", 4 * 1024 * 1024 * 1024)
    task_log_id = task_log.id  # capture for reliable re-fetch after any rollback

    def set_meta(**updates):
        task_log.meta = {**(task_log.meta or {}), **updates}

    try:
        zf = zipfile.ZipFile(zip_path)
    except zipfile.BadZipFile as e:
        raise ValueError(f"无法读取 ZIP 文件: {str(e)}")

    with zf:
        bad = zf.testzip()
        if bad is not None:
            raise ValueError(f"ZIP 文件损坏: {bad}")

        infos = zf.infolist()
        if len(infos) > max_entries:
            raise ValueError(f"ZIP 包含超过 {max_entries} 个文件")

        total_uncompressed = sum(zi.file_size for zi in infos)
        if total_uncompressed > max_extracted:
            raise ValueError("ZIP 解压后总体积超过限制")

        # Path-traversal / symlink safety (we never extract to disk, but reject anyway).
        for zi in infos:
            name = zi.filename
            if name.startswith("/") or "\\" in name or ".." in Path(name).parts:
                raise ValueError(f"ZIP 包含非法路径: {name}")
            mode = (zi.external_attr >> 16) & 0o170000
            if mode == 0o120000:  # symlink
                raise ValueError(f"ZIP 包含符号链接: {name}")

        excel_infos = [
            zi for zi in infos
            if not zi.is_dir() and zi.filename.lower().endswith((".xlsx", ".xls"))
        ]
        if not excel_infos:
            raise ValueError("ZIP 中未找到 Excel 文件 (.xlsx/.xls)")
        warnings = []
        excel_names = [_decode_zip_name(zi) for zi in excel_infos]
        excel_zi = (
            excel_infos[excel_names.index("菜品导入模板.xlsx")]
            if "菜品导入模板.xlsx" in excel_names
            else excel_infos[0]
        )
        excel_name = _decode_zip_name(excel_zi)
        if len(excel_infos) > 1:
            warnings.append(f"ZIP 中找到多个 Excel 文件，已使用「{excel_name}」")

        set_meta(status_text="解析 Excel 中")
        db.session.commit()

        try:
            df = pd.read_excel(io.BytesIO(zf.read(excel_zi)), sheet_name=0, dtype=str)
            df = df.fillna("").map(lambda x: str(x).strip() if x else "")
        except Exception as e:
            logger.error("Failed to parse Excel from zip: %s", e)
            raise ValueError(f"解析 Excel 失败: {str(e)}")

        set_meta(status_text="导入菜品信息中")
        db.session.commit()

        created, updated, errors, dish_map = _upsert_dishes_from_dataframe(df)
        if errors and not created and not updated:
            raise ValueError("导入失败:\n" + "\n".join(errors[:20]))

        db.session.flush()  # populate ids on newly created dishes
        db.session.commit()  # persist dishes so ids are stable for image binding

        # Group sample images by top-level folder. Only "<folder>/<file>"
        # entries are honored; deeper nesting and macOS metadata are skipped.
        # Names are re-decoded via _decode_zip_name so macOS UTF-8-without-flag
        # archives match dish names; image bytes are read via the ZipInfo object
        # to bypass the (mojibake) NameToInfo lookup.
        folder_images = {}
        for zi in infos:
            if zi.is_dir():
                continue
            name = _decode_zip_name(zi)
            parts = name.split("/")
            if len(parts) > 2:
                # Skip macOS resource-fork noise silently; warn on other nesting.
                if parts[0] != "__MACOSX":
                    warnings.append(f"忽略嵌套目录文件: {name}")
                continue
            if len(parts) != 2 or not parts[0] or not parts[1]:
                continue
            folder, basename = parts[0], parts[1]
            if folder == "__MACOSX" or basename.startswith("."):
                continue
            ext = os.path.splitext(basename)[1].lower()
            if ext not in ALLOWED_SAMPLE_IMAGE_EXTENSIONS:
                continue
            folder_images.setdefault(folder, []).append((basename, zi))

        folder_total = len(folder_images)
        task_log.total_count = folder_total
        set_meta(total_count=folder_total, processed_count=0, status_text=f"处理图片 0/{folder_total}")
        db.session.commit()

        images_imported = 0
        images_skipped = 0
        dishes_with_images = 0
        folders_unmatched = 0
        processed_folders = 0

        for folder, entries in folder_images.items():
            status_label = folder
            try:
                dish = dish_map.get(_normalize_dish_name(folder))
                if dish is None:
                    folders_unmatched += 1
                    images_skipped += len(entries)
                    warnings.append(f"未找到与文件夹「{folder}」匹配的菜品")
                else:
                    status_label = dish.name
                    entries_sorted = sorted(entries, key=lambda t: t[0])
                    existing_active = DishSampleImage.query.filter_by(
                        dish_id=dish.id, is_active=True,
                    ).count()
                    if existing_active >= MAX_DISH_SAMPLE_IMAGES:
                        images_skipped += len(entries_sorted)
                        warnings.append(
                            f"菜品「{dish.name}」样图已达上限({MAX_DISH_SAMPLE_IMAGES})，跳过 {len(entries_sorted)} 张"
                        )
                    else:
                        to_add = min(MAX_DISH_SAMPLE_IMAGES - existing_active, len(entries_sorted))
                        current_max_sort = db.session.query(db.func.max(DishSampleImage.sort_order)).filter(
                            DishSampleImage.dish_id == dish.id,
                            DishSampleImage.is_active.is_(True),
                        ).scalar() or 0
                        has_cover = db.session.query(DishSampleImage.id).filter(
                            DishSampleImage.dish_id == dish.id,
                            DishSampleImage.is_cover.is_(True),
                            DishSampleImage.is_active.is_(True),
                        ).first()

                        added_for_dish = 0
                        for basename, zi_img in entries_sorted:
                            if added_for_dish >= to_add:
                                images_skipped += 1
                                continue
                            try:
                                data = zf.read(zi_img)
                            except Exception as e:  # pragma: no cover - defensive
                                warnings.append(f"读取 {basename} 失败: {str(e)}")
                                images_skipped += 1
                                continue
                            magic_err = _validate_image_magic_bytes(data)
                            if magic_err:
                                warnings.append(f"{basename}: {magic_err}")
                                images_skipped += 1
                                continue
                            adapter = _BytesFileStorage(basename, data)
                            val_err = _validate_sample_image_file(adapter)
                            if val_err:
                                warnings.append(f"{basename}: {val_err}")
                                images_skipped += 1
                                continue
                            current_max_sort += 1
                            image = _save_sample_image_file(
                                dish.id,
                                adapter,
                                sort_order=current_max_sort,
                                is_cover=(added_for_dish == 0 and not has_cover),
                            )
                            db.session.add(image)
                            images_imported += 1
                            added_for_dish += 1
                        if added_for_dish > 0:
                            dishes_with_images += 1

                processed_folders += 1
                set_meta(
                    processed_count=processed_folders,
                    current_dish_name=status_label,
                    status_text=f"处理图片 {processed_folders}/{folder_total}：{status_label}",
                )
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                task_log = TaskLog.query.get(task_log_id)  # re-fetch after rollback
                logger.error("Failed to process folder %s in zip import: %s", folder, e, exc_info=True)
                warnings.append(f"处理文件夹「{folder}」失败: {str(e)}")
                processed_folders += 1
                if task_log:
                    set_meta(processed_count=processed_folders, status_text=f"处理图片 {processed_folders}/{folder_total}")
                    db.session.commit()

    combined_warnings = (errors + warnings)[:10]
    return {
        "created_count": len(created),
        "updated_count": len(updated),
        "created": created[:20],
        "updated": updated[:20],
        "images_imported": images_imported,
        "images_skipped": images_skipped,
        "dishes_with_images": dishes_with_images,
        "folders_unmatched": folders_unmatched,
        "warnings": combined_warnings,
    }


@bp.route("/analyze-nutrition-preview", methods=["POST"])
@role_required(*ALLOWED_ROLES_WRITE)
def preview_dish_nutrition():
    """Preview nutrition analysis for a dish name without saving."""
    data = request.get_json() or {}
    dish_name = data.get("dish_name", "").strip()
    weight = int(data.get("weight", 100))
    ingredients = data.get("ingredients", "").strip()

    if not dish_name:
        return api_error("菜品名称不能为空")

    if weight <= 0 or weight > 10000:
        return api_error("重量必须在 1-10000g 之间")

    config = current_app.config
    api_key = config.get("OPENAI_API_KEY", "")

    if not api_key:
        return api_error("营养分析服务未配置 (OPENAI_API_KEY)"), 503

    try:
        analyzer = DishAnalyzerService(config)
        result = analyzer.analyze_nutrition(dish_name, weight, ingredients)

        return api_ok({
            "dish_name": dish_name,
            "weight": weight,
            "category": result.get("category", ""),
            "nutrition": {
                "calories": result.get("calories"),
                "protein": result.get("protein"),
                "fat": result.get("fat"),
                "carbohydrate": result.get("carbohydrate"),
                "sodium": result.get("sodium"),
                "fiber": result.get("fiber"),
            },
            "description": result.get("description", ""),
            "structured_description": result.get("structured_description", {}),
            "notes": result.get("notes", ""),
        })
    except Exception as e:
        logger.error(f"Failed to preview dish nutrition: {e}")
        return api_error(f"营养分析失败: {str(e)}"), 500


@bp.route("/batch-analyze-nutrition", methods=["POST"])
@role_required(*ALLOWED_ROLES_WRITE)
def batch_analyze_nutrition():
    """Queue nutrition analysis for all dishes without nutrition data."""
    config = current_app.config
    api_key = config.get("OPENAI_API_KEY", "")

    if not api_key:
        return api_error("营养分析服务未配置 (OPENAI_API_KEY)"), 503

    active_task = TaskLog.query.filter(
        TaskLog.task_type == "dish_nutrition_analysis",
        TaskLog.status.in_(("pending", "running")),
    ).order_by(TaskLog.started_at.desc()).first()
    if active_task:
        return api_ok({
            "message": "已有批量营养分析任务正在执行",
            "task_id": active_task.id,
            "task": active_task.to_dict(),
        })

    total = Dish.query.filter(
        Dish.is_active.is_(True),
        Dish.calories.is_(None),
    ).count()

    if total == 0:
        return api_ok({"message": "没有需要分析的菜品", "total": 0, "success": 0, "failed": 0, "errors": []})

    from app.tasks.dishes import batch_analyze_dish_nutrition, create_batch_nutrition_task_log

    task_log = create_batch_nutrition_task_log(total)
    try:
        celery_task = batch_analyze_dish_nutrition.delay(task_log.id)
        task_log.meta = {
            **(task_log.meta or {}),
            "celery_task_id": celery_task.id,
            "status_text": "任务已提交，等待执行",
        }
        db.session.commit()
    except Exception as e:
        logger.error("Failed to submit dish nutrition analysis task: %s", e, exc_info=True)
        task_log.status = "failed"
        task_log.error_count = 1
        task_log.error_message = str(e)
        task_log.finished_at = datetime.utcnow()
        task_log.meta = {
            **(task_log.meta or {}),
            "status_text": "任务提交失败",
        }
        db.session.commit()
        return api_error(f"提交批量分析任务失败: {str(e)}"), 500

    return api_ok({
        "message": "批量营养分析任务已提交",
        "task_id": task_log.id,
        "task": task_log.to_dict(),
        "total": total,
    })


@bp.route("/generate-description", methods=["POST"])
@role_required(*ALLOWED_ROLES_WRITE)
def generate_dish_description():
    """Generate visual description for a dish from an uploaded sample image using VL model."""
    if "image" not in request.files:
        return api_error("请上传图片文件")

    file = request.files["image"]
    if not file.filename:
        return api_error("文件名无效")

    # Check file extension
    allowed_extensions = {".jpg", ".jpeg", ".png", ".webp"}
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in allowed_extensions:
        return api_error(f"不支持的图片格式，请上传 {', '.join(allowed_extensions)} 格式")

    # Get optional dish name for context
    dish_name = request.form.get("dish_name", "").strip()

    config = current_app.config
    api_key = config.get("QWEN_API_KEY", "")
    if not api_key:
        return api_error("VL服务未配置 (QWEN_API_KEY)"), 503

    # Save to temp file and process
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name

        vl_service = QwenVLService(config)
        result = vl_service.describe_dishes(tmp_path)

        # Clean up temp file
        os.unlink(tmp_path)

        description = result.get("description", "")
        if not description:
            return api_error("无法从图片生成描述")

        descriptions = result.get("descriptions", [])
        # If dish name provided, prepend it for context
        if dish_name:
            description = f"【{dish_name}】{description}"
            if descriptions:
                descriptions = [dict(descriptions[0], description=description), *descriptions[1:]]

        return api_ok({
            "description": description,
            "structured_description": result.get("structured_description", empty_structured_description()),
            "notes": result.get("notes", ""),
            "descriptions": descriptions,
            "dish_name": dish_name,
        })
    except Exception as e:
        logger.error(f"Failed to generate dish description: {e}")
        # Clean up temp file on error
        if "tmp_path" in locals():
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
        return api_error(f"生成描述失败: {str(e)}"), 500
